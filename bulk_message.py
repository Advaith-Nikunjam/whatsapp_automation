import pywhatkit
from openpyxl import workbook,load_workbook
import time
import random
def bulk_message(message_text, excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    for i in range(1,ws.max_row+1):
        number = str(int(i))
        number = ws[f'A{i}'].value
        phone_number = f'+91{number}'
        #print(phone_number)

        pywhatkit.sendwhatmsg_instantly(

            phone_no = phone_number,
            message = message_text,
            wait_time = random.randint(35,45),
            tab_close = True

            )
        time.sleep(random.randint(20,40))